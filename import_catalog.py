"""
Скрипт для импорта каталога оборудования из Excel в JSON

Читает файл PriceCatalogs/Каталог оборудования.xlsx
Генерирует public/data/catalog.json для использования в приложении
"""

import openpyxl
import json
import os
import re
from pathlib import Path

def is_category_header(row, ws):
    """Определяет, является ли строка заголовком категории"""
    # Категории обычно имеют заполненную только первую ячейку
    # и могут быть выделены цветом (синие)
    cell_a = ws.cell(row=row, column=1)
    cell_b = ws.cell(row=row, column=2)
    cell_c = ws.cell(row=row, column=3)
    
    # Если в первой колонке есть значение, а во второй и третьей пусто
    if cell_a.value and not cell_b.value and not cell_c.value:
        # Проверяем что это не просто пустая строка
        value = str(cell_a.value).strip()
        if value:
            # Дополнительно можно проверить цвет заливки
            # Синие заголовки обычно имеют fill
            return True
    return False

def is_subcategory(row_value):
    """Определяет, является ли строка подкатегорией (производитель)"""
    if not row_value:
        return False
    
    value = str(row_value).strip()
    # Подкатегории обычно имеют отступы или особый формат
    # Например: "               AM" (с пробелами)
    if value and (value.startswith(' ' * 5) or len(value) < 20):
        return True
    return False

def clean_price(price_value):
    """Очищает и конвертирует цену в число"""
    if price_value is None:
        return 0
    
    if isinstance(price_value, (int, float)):
        return float(price_value)
    
    # Если строка, убираем все кроме цифр и точки/запятой
    price_str = str(price_value).replace(',', '.').replace(' ', '')
    try:
        return float(price_str)
    except:
        return 0

def parse_catalog_excel(excel_path):
    """Парсит Excel файл и возвращает структурированные данные"""
    print(f"📖 Открываем файл: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    catalog_data = {
        "categories": [],
        "items": []
    }
    
    current_category = None
    current_subcategory = None
    item_id = 1
    
    print(f"📊 Обрабатываем {ws.max_row} строк...")
    
    for row_num in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_num, column=1).value
        cell_b = ws.cell(row=row_num, column=2).value
        cell_c = ws.cell(row=row_num, column=3).value
        
        # Пропускаем полностью пустые строки
        if not any([cell_a, cell_b, cell_c]):
            continue
        
        # Проверяем на заголовок категории
        if is_category_header(row_num, ws):
            current_category = str(cell_a).strip()
            current_subcategory = None
            
            if current_category not in catalog_data["categories"]:
                catalog_data["categories"].append(current_category)
                print(f"📁 Найдена категория: {current_category}")
            continue
        
        # Проверяем на подкатегорию
        if cell_a and not cell_b and not cell_c:
            if is_subcategory(cell_a):
                current_subcategory = str(cell_a).strip()
                print(f"  📂 Подкатегория: {current_subcategory}")
                continue
        
        # Это товарная позиция (есть артикул, название и цена)
        if cell_a and cell_b:
            article = str(cell_a).strip()
            name = str(cell_b).strip()
            price = clean_price(cell_c)
            
            # Создаем товар
            item = {
                "id": item_id,
                "article": article,
                "name": name,
                "price": price,
                "category": current_category or "Без категории",
                "subcategory": current_subcategory or ""
            }
            
            catalog_data["items"].append(item)
            item_id += 1
    
    wb.close()
    
    print(f"\n✅ Обработка завершена:")
    print(f"   📂 Категорий: {len(catalog_data['categories'])}")
    print(f"   📦 Товаров: {len(catalog_data['items'])}")
    
    return catalog_data

def save_catalog_json(catalog_data, output_path):
    """Сохраняет данные каталога в JSON файл"""
    # Создаем директорию если её нет
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    print(f"\n💾 Сохраняем в файл: {output_path}")
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(catalog_data, f, ensure_ascii=False, indent=2)
    
    # Статистика файла
    file_size = os.path.getsize(output_path)
    size_mb = file_size / (1024 * 1024)
    
    print(f"✅ Файл сохранен!")
    print(f"   📊 Размер: {size_mb:.2f} MB")
    print(f"   📁 Путь: {output_path}")

def main():
    """Главная функция"""
    print("=" * 80)
    print("🔧 ИМПОРТ КАТАЛОГА ОБОРУДОВАНИЯ")
    print("=" * 80)
    
    # Пути
    excel_path = "PriceCatalogs/Каталог оборудования.xlsx"
    json_path = "public/data/catalog.json"
    
    # Проверяем существование Excel файла
    if not os.path.exists(excel_path):
        print(f"❌ Ошибка: файл {excel_path} не найден!")
        return
    
    try:
        # Парсим Excel
        catalog_data = parse_catalog_excel(excel_path)
        
        # Сохраняем JSON
        save_catalog_json(catalog_data, json_path)
        
        print("\n" + "=" * 80)
        print("🎉 ИМПОРТ УСПЕШНО ЗАВЕРШЕН!")
        print("=" * 80)
        
        # Показываем примеры
        print("\n📋 Примеры товаров из каталога:")
        for i, item in enumerate(catalog_data["items"][:5], 1):
            print(f"{i}. [{item['article']}] {item['name'][:60]}... - {item['price']:,.0f} ₽")
            print(f"   Категория: {item['category']} / {item['subcategory']}")
        
    except Exception as e:
        print(f"\n❌ Ошибка при импорте: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
