import openpyxl
import sys

try:
    # Load the workbook (without read_only to access dimensions)
    wb = openpyxl.load_workbook('PriceCatalogs/Каталог оборудования.xlsx')
    
    print('=' * 80)
    print('СТРУКТУРА ФАЙЛА "Каталог оборудования.xlsx"')
    print('=' * 80)
    
    print(f'\nЛисты в файле: {wb.sheetnames}')
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n{"=" * 80}')
        print(f'ЛИСТ: {sheet_name}')
        print(f'{"=" * 80}')
        print(f'Размеры: {ws.dimensions}')
        
        # Get headers
        print('\nПервые 20 строк:')
        print('-' * 80)
        
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            # Clean up the row - remove None values at the end
            row_data = list(row)
            while row_data and row_data[-1] is None:
                row_data.pop()
            
            if row_data:  # Only print non-empty rows
                print(f'{i:3d}. {row_data}')
            
            if i >= 20:
                break
        
        # Count total rows with data
        total_rows = 0
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                total_rows += 1
        
        print(f'\nВсего строк с данными: {total_rows}')
    
    wb.close()
    print('\n' + '=' * 80)
    print('АНАЛИЗ ЗАВЕРШЕН')
    print('=' * 80)

except Exception as e:
    print(f'Ошибка: {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)
